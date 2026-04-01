"""
Modul pro export dat z dashboardu do XLSX, CSV, PNG a PDF.
"""

import io
import os
import tempfile
import zipfile
from datetime import date, datetime

import pandas as pd
import streamlit as st


# ---------------------------------------------------------------------------
# Pomocné funkce – ošetření tz-aware datetime indexu pro Excel kompatibilitu
# ---------------------------------------------------------------------------

def _tz_strip(df: pd.DataFrame) -> pd.DataFrame:
    """Vrátí kopii DataFrame s tz-naive indexem (pro Excel/CSV kompatibilitu)."""
    df = df.copy()
    if hasattr(df.index, "tz") and df.index.tz is not None:
        df.index = df.index.tz_localize(None)
    return df


# ---------------------------------------------------------------------------
# Part 1: XLSX
# ---------------------------------------------------------------------------

def _generuj_xlsx(datasety: dict[str, pd.DataFrame]) -> bytes:
    """Generuje XLSX soubor s jedním listem na dataset. Vrací bytes."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for nazev, df in datasety.items():
            sheet_name = nazev[:31]  # Excel max 31 znaků
            df_clean = _tz_strip(df)
            df_clean.to_excel(writer, sheet_name=sheet_name)

            # Tučné hlavičky + formátování datumů
            ws = writer.sheets[sheet_name]
            from openpyxl.styles import Font
            bold = Font(bold=True)
            for cell in ws[1]:
                cell.font = bold
            # Český formát datumů pro datetime sloupce
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (pd.Timestamp, datetime)):
                        cell.number_format = "DD.MM.YYYY HH:MM"

    return buf.getvalue()


# ---------------------------------------------------------------------------
# Part 2: CSV
# ---------------------------------------------------------------------------

def _generuj_csv(datasety: dict[str, pd.DataFrame]) -> tuple[bytes, str]:
    """
    Jeden dataset → CSV (ext 'csv').
    Více datasetů → ZIP s CSV soubory (ext 'zip').
    UTF-8 s BOM pro správné české znaky v Excelu.
    """
    if len(datasety) == 1:
        nazev, df = next(iter(datasety.items()))
        df_clean = _tz_strip(df)
        csv_bytes = df_clean.to_csv().encode("utf-8-sig")
        return csv_bytes, "csv"

    # Více datasetů – ZIP
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for nazev, df in datasety.items():
            df_clean = _tz_strip(df)
            csv_bytes = df_clean.to_csv().encode("utf-8-sig")
            zf.writestr(f"{nazev}.csv", csv_bytes)
    return buf.getvalue(), "zip"


# ---------------------------------------------------------------------------
# Part 3: PNG
# ---------------------------------------------------------------------------

def _generuj_png(grafy: list, nazvy_grafu: list[str]) -> tuple[bytes, str]:
    """
    Jeden graf → PNG (ext 'png').
    Více grafů → ZIP s PNG soubory (ext 'zip').
    """
    if len(grafy) == 1:
        img_bytes = grafy[0].to_image(
            format="png", width=1200, height=600, scale=2
        )
        return img_bytes, "png"

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fig, nazev in zip(grafy, nazvy_grafu):
            img_bytes = fig.to_image(
                format="png", width=1200, height=600, scale=2
            )
            safe_name = nazev.replace("/", "_").replace("\\", "_")
            zf.writestr(f"{safe_name}.png", img_bytes)
    return buf.getvalue(), "zip"


# ---------------------------------------------------------------------------
# Part 4: PDF
# ---------------------------------------------------------------------------

def _generuj_pdf(
    grafy: list,
    nazvy_grafu: list[str],
    nazev_stranky: str,
    datum_od: date,
    datum_do: date,
) -> bytes:
    """
    Landscape A4 PDF – titulní strana + jedna strana na graf.
    Používá fpdf2 s vestavěným Helvetica fontem.
    """
    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=False)

    # --- Titulní strana ---
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 28)
    pdf.ln(50)
    pdf.cell(0, 15, nazev_stranky, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)
    pdf.set_font("Helvetica", "", 16)
    pdf.cell(
        0, 10,
        f"{datum_od.strftime('%d.%m.%Y')} - {datum_do.strftime('%d.%m.%Y')}",
        align="C", new_x="LMARGIN", new_y="NEXT",
    )
    pdf.ln(8)
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(
        0, 10,
        f"Exportovano: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        align="C", new_x="LMARGIN", new_y="NEXT",
    )

    # --- Stránky s grafy ---
    with tempfile.TemporaryDirectory() as tmpdir:
        for i, (fig, nazev) in enumerate(zip(grafy, nazvy_grafu)):
            img_path = os.path.join(tmpdir, f"chart_{i}.png")
            fig.write_image(img_path, width=1200, height=600, scale=2)

            pdf.add_page()
            pdf.set_font("Helvetica", "B", 16)
            pdf.cell(0, 12, nazev, align="C", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(3)

            # A4 landscape: 297 x 210 mm, margins ~10mm
            img_w = 277  # šířka obrázku v mm
            img_h = 140  # výška obrázku v mm
            x = (297 - img_w) / 2
            y = pdf.get_y()
            pdf.image(img_path, x=x, y=y, w=img_w, h=img_h)

    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# Part 5: Sidebar – render_export_sidebar
# ---------------------------------------------------------------------------

def render_export_sidebar(
    nazev_stranky: str,
    filtrovana_data: dict[str, pd.DataFrame],
    surova_data_soubory: dict[str, str],
    grafy: list,
    nazvy_grafu: list[str],
    datum_od: date,
    datum_do: date,
    cache_slozka: str | None = None,
):
    """
    Vykreslí export sekci v sidebaru.

    Parametry:
        nazev_stranky:      název stránky (pro klíče a název souboru)
        filtrovana_data:    dict {název: DataFrame} s filtrovanými daty
        surova_data_soubory: dict {název: cesta_k_parquet} pro surová data
        grafy:              list Plotly figur
        nazvy_grafu:        list názvů grafů
        datum_od:           počáteční datum
        datum_do:           koncové datum
        cache_slozka:       cesta ke složce s cache (pro surová data)
    """
    prefix = nazev_stranky.replace(" ", "_")

    if cache_slozka is None:
        cache_slozka = os.path.join(os.path.dirname(__file__), "data", "cache")

    with st.sidebar:
        st.divider()
        st.caption("📥 Export dat")

        # --- Typ dat ---
        typ_dat = st.radio(
            "Typ dat",
            ["Filtrovaná data (aktuální pohled)", "Surová data (kompletní dataset)"],
            key=f"{prefix}_export_typ",
        )

        je_filtrovana = typ_dat == "Filtrovaná data (aktuální pohled)"

        # --- Výběr datasetů ---
        if je_filtrovana:
            datasety_k_exportu = filtrovana_data
        else:
            dostupne = list(surova_data_soubory.keys())
            vybrane = st.multiselect(
                "Datasety k exportu",
                dostupne,
                default=dostupne,
                key=f"{prefix}_export_datasets",
            )
            if not vybrane:
                st.warning("Vyber alespoň jeden dataset.")
                return

            # Ověř existenci souborů, načtení odloženo do generování
            surove_cesty = {}
            for nazev in vybrane:
                cesta = surova_data_soubory[nazev]
                if not os.path.isabs(cesta):
                    cesta = os.path.join(cache_slozka, cesta)
                if os.path.exists(cesta):
                    surove_cesty[nazev] = cesta

            if not surove_cesty:
                st.warning("Žádná data k exportu.")
                return
            datasety_k_exportu = None  # marker: data se načtou až při generování

        # --- Rozsah dat ---
        st.caption("Rozsah exportu:")
        exp_od = st.date_input(
            "Export od",
            value=datum_od,
            min_value=datum_od,
            max_value=datum_do,
            key=f"{prefix}_export_od",
        )
        exp_do = st.date_input(
            "Export do",
            value=datum_do,
            min_value=datum_od,
            max_value=datum_do,
            key=f"{prefix}_export_do",
        )

        if exp_od > exp_do:
            st.error("Datum Od musí být před datem Do.")
            return

        # --- Formát ---
        formaty = ["XLSX", "CSV"]
        if je_filtrovana:
            formaty.extend(["PDF", "PNG"])

        format_export = st.radio(
            "Formát exportu",
            formaty,
            key=f"{prefix}_export_format",
        )

        # --- Pomocná funkce: načti a filtruj datasety ---
        def _priprav_datasety() -> dict[str, pd.DataFrame]:
            ts_od = pd.Timestamp(exp_od)
            ts_do = pd.Timestamp(exp_do) + pd.Timedelta(days=1)

            if je_filtrovana:
                zdroj = datasety_k_exportu
            else:
                # Deferred loading surových dat
                zdroj = {}
                for nazev, cesta in surove_cesty.items():
                    zdroj[nazev] = pd.read_parquet(cesta)

            vysledek = {}
            for nazev, df in zdroj.items():
                idx = df.index
                # Non-datetime index (e.g. string) → include as-is
                if not isinstance(idx, pd.DatetimeIndex):
                    if len(df) > 0:
                        vysledek[nazev] = df
                    continue
                if hasattr(idx, "tz") and idx.tz is not None:
                    ts_od_cmp = ts_od.tz_localize(str(idx.tz)) if ts_od.tz is None else ts_od
                    ts_do_cmp = ts_do.tz_localize(str(idx.tz)) if ts_do.tz is None else ts_do
                else:
                    ts_od_cmp = ts_od.tz_localize(None) if ts_od.tz is not None else ts_od
                    ts_do_cmp = ts_do.tz_localize(None) if ts_do.tz is not None else ts_do
                mask = (idx >= ts_od_cmp) & (idx < ts_do_cmp)
                df_slice = df[mask]
                if len(df_slice) > 0:
                    vysledek[nazev] = df_slice
            return vysledek

        # --- Generování a stažení ---
        datum_od_str = exp_od.strftime("%Y%m%d")
        datum_do_str = exp_do.strftime("%Y%m%d")
        base_name = f"{nazev_stranky}_{datum_od_str}_{datum_do_str}"

        if format_export == "XLSX":
            datasety_final = _priprav_datasety()
            if not datasety_final:
                st.warning("Žádná data ve zvoleném rozsahu.")
                return
            data_bytes = _generuj_xlsx(datasety_final)
            st.download_button(
                "⬇️ Stáhnout XLSX",
                data=data_bytes,
                file_name=f"{base_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"{prefix}_dl_xlsx",
            )

        elif format_export == "CSV":
            datasety_final = _priprav_datasety()
            if not datasety_final:
                st.warning("Žádná data ve zvoleném rozsahu.")
                return
            data_bytes, ext = _generuj_csv(datasety_final)
            mime = "text/csv" if ext == "csv" else "application/zip"
            st.download_button(
                "⬇️ Stáhnout CSV" if ext == "csv" else "⬇️ Stáhnout ZIP (CSV)",
                data=data_bytes,
                file_name=f"{base_name}.{ext}",
                mime=mime,
                key=f"{prefix}_dl_csv",
            )

        elif format_export == "PNG":
            if not grafy:
                st.warning("Žádné grafy k exportu.")
                return
            data_bytes, ext = _generuj_png(grafy, nazvy_grafu)
            mime = "image/png" if ext == "png" else "application/zip"
            st.download_button(
                "⬇️ Stáhnout PNG" if ext == "png" else "⬇️ Stáhnout ZIP (PNG)",
                data=data_bytes,
                file_name=f"{base_name}.{ext}",
                mime=mime,
                key=f"{prefix}_dl_png",
            )

        elif format_export == "PDF":
            if not grafy:
                st.warning("Žádné grafy k exportu.")
                return
            data_bytes = _generuj_pdf(
                grafy, nazvy_grafu, nazev_stranky, exp_od, exp_do
            )
            st.download_button(
                "⬇️ Stáhnout PDF",
                data=data_bytes,
                file_name=f"{base_name}.pdf",
                mime="application/pdf",
                key=f"{prefix}_dl_pdf",
            )
